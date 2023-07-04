import os
import tkinter as tk
from tkinter import filedialog
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def seleccionar_archivo():
    archivo = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    entrada_archivo.set(archivo)

def calcular_diferencias(historial):
    diferencias = [None]  
    for i in range(1, len(historial)):
        diferencia = historial.iloc[i] - historial.iloc[i-1]  
        diferencias.append(diferencia)
    return diferencias

def ejecutar_proceso():
    archivo_excel = entrada_archivo.get()

    if archivo_excel:
        
        df = pd.read_excel(archivo_excel)

        
        df.sort_values(['Medidor', 'Fecha'], inplace=True)

        
        df_ordenado = df.copy()

        
        df_ordenado['Diferencias_Activa'] = df_ordenado.groupby('Medidor')['Activa'].transform(calcular_diferencias)
        df_ordenado['Diferencias_Reactiva'] = df_ordenado.groupby('Medidor')['Reactiva'].transform(calcular_diferencias)

        
        columnas_ordenadas = ['T', 'S', 'P', 'R', 'Recorrido', 'Medidor', 'Fecha', 'Activa', 'Reactiva', 'Diferencias_Activa', 'Diferencias_Reactiva', 'Nº Cliente', 'Domicilio', 'Leer Reactiva', 'Asignado A', 'Nº Legajo', 'Nombre', 'Anomalía']
        df_ordenado = df_ordenado[columnas_ordenadas]

        
        carpeta_programa = os.path.dirname(os.path.abspath(__file__))

        
        nuevo_archivo_excel_ordenado = os.path.join(carpeta_programa, 'nuevo_archivo_excel_ordenado.xlsx')
        df_ordenado.to_excel(nuevo_archivo_excel_ordenado, index=False)

       
        wb = load_workbook(nuevo_archivo_excel_ordenado)
        ws = wb.active
        fill_red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        for col in ws.iter_cols(min_col=9, max_col=10):
            for cell in col:
                if cell.column in [9, 10]:
                    value = cell.value
                    if isinstance(value, (int, float)) and (value < 0 or value > 4000):
                        cell.fill = fill_red

       
        nuevo_archivo_excel_final = os.path.join(carpeta_programa, 'nuevo_archivo_excel_final.xlsx')
        wb.save(nuevo_archivo_excel_final)

        resultado.set("Proceso completado. Archivo guardado como: " + nuevo_archivo_excel_final)
    else:
        resultado.set("Error: No se ha seleccionado ningún archivo")


ventana = tk.Tk()
ventana.title("Procesamiento de Archivo Excel")
ventana.geometry("400x200")


entrada_archivo = tk.StringVar()


btn_seleccionar = tk.Button(ventana, text="Seleccionar Archivo", command=seleccionar_archivo)
btn_seleccionar.pack(pady=10)

lbl_ruta_archivo = tk.Label(ventana, textvariable=entrada_archivo)
lbl_ruta_archivo.pack()

btn_procesar = tk.Button(ventana, text="Iniciar Proceso", command=ejecutar_proceso)
btn_procesar.pack(pady=10)

resultado = tk.StringVar()

lbl_resultado = tk.Label(ventana, textvariable=resultado)
lbl_resultado.pack()


ventana.mainloop()
